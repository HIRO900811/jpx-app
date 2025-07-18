import requests
from bs4 import BeautifulSoup
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# JPXの建玉ページURL
JPX_URL = "https://www.jpx.co.jp/markets/derivatives/trading-volume/index.html"

# 最新ExcelファイルのURLを取得
def get_latest_excel_url():
    res = requests.get(JPX_URL)
    soup = BeautifulSoup(res.text, 'html.parser')
    for a in soup.find_all('a', href=True):
        href = a['href']
        if href.endswith(".xlsx") and "interest" in href.lower():
            return "https://www.jpx.co.jp" + href if href.startswith("/") else href
    return None

# Excelの中身をHTMLに変換して返す
def parse_excel_to_html():
    url = get_latest_excel_url()
    if not url:
        return "<p>❌ Excelファイルが見つかりませんでした。</p>"

    res = requests.get(url)
    excel_file = BytesIO(res.content)
    df_dict = pd.read_excel(excel_file, sheet_name=None)

    html = ""
    for sheet_name, df in df_dict.items():
        df.fillna(method='ffill', inplace=True)
        html += f"<h2>{sheet_name}</h2>\n"
        html += df.to_html(classes="table table-bordered", index=False)
    return html

# Excelから特定のセルを抽出
def extract_specific_cells_from_excel(xlsx_bytes):
    workbook = load_workbook(BytesIO(xlsx_bytes), data_only=True)

    # 🔍 シート名一覧を表示
    print("📄 シート一覧:", workbook.sheetnames)

    sheet_name = "デリバティブ建玉残高状況"
    if sheet_name not in workbook.sheetnames:
        return f"❌ シート「{sheet_name}」が見つかりません"

    ws = workbook[sheet_name]

    # セルを辞書にまとめて返す
    cells = {
        "nikkei225": ws["B30"].value,
        "nikkei225_diff": ws["E30"].value,
        "nikkei225_total_diff": ws["E49"].value,
        "nikkei225mini_total": ws["L52"].value,
        "topix": ws["B50"].value,
        "topix_diff": ws["E50"].value,
        "put_volume": ws["C295"].value,
        "put_diff": ws["E295"].value,
        "call_volume": ws["C296"].value,
        "call_diff": ws["E296"].value,
    }

    return cells

#JPXのリアルタイムデータAPIから「日経225」の先物データを取得する
def fetch_nikkei225_from_api():
    url = "https://port.jpx.co.jp/jpxhp/jcgi/wrap/qjsonp.aspx?F=ctl/future&DISPTYPE=daytime"
    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    try:
        res = requests.get(url, headers=headers, timeout=5)
        res.encoding = res.apparent_encoding
        data = res.json()

        if data["section1"]["status"] == 0 and data["section1"]["hitcount"] > 0:
            rows = data["section1"]["data"]

            for item in rows:
                if "日経225" in item.get("name", ""):
                    for future in item["future"]:
                        return {
                            "name": item["name"],
                            "limit_month": future["DELI"],
                            "price": future["DPP"],
                            "diff": future["DYWP"]
                        }
        return {"error": "データが見つかりませんでした"}

    except Exception as e:
        return {"error": f"APIエラー: {str(e)}"}

#出力データの分析
def analyze_market_trend(nikkei_data, extracted_data):
    result = {
        "futures_trend": "",
        "options_trend": "",
    }

    # --- 共通: 株価前日比（+/-）
    price_change = nikkei_data.get("diff", "")
    is_price_up = price_change.startswith("+")
    is_price_down = price_change.startswith("-")

    # --- 先物建玉動向 ---
    try:
        nikkei225_diff = int(extracted_data.get("nikkei225_diff", 0))
        nikkei225_total_diff = int(extracted_data.get("nikkei225_total_diff", 0))
        is_futures_increase = nikkei225_diff > 0 and nikkei225_total_diff > 0
        is_futures_decrease = nikkei225_diff < 0 and nikkei225_total_diff < 0

        if is_price_up and is_futures_increase:
            result["futures_trend"] = "📈 上昇方向のトレンド発生を意識"
        elif is_price_up and is_futures_decrease:
            result["futures_trend"] = "⚠ 上昇は長続きせず反転下落の可能性"
        elif is_price_down and is_futures_increase:
            result["futures_trend"] = "🔄 下落は長続きせず再上昇の可能性"
        elif is_price_down and is_futures_decrease:
            result["futures_trend"] = "📉 下落方向のトレンド発生を意識"
        else:
            result["futures_trend"] = "❔ 判断困難"

    except Exception as e:
        result["futures_trend"] = f"❌ エラー: {e}"

    # --- オプション建玉動向 ---
    try:
        call_diff = int(extracted_data.get("call_diff", 0))
        put_diff = int(extracted_data.get("put_diff", 0))

        if is_price_up and call_diff > 0:
            result["options_trend"] = "💪 強気（上昇＆コール建玉増）"
        elif is_price_up and call_diff < 0:
            result["options_trend"] = "📉 上昇は長く続きにくい（コール減少）"
        elif is_price_down and put_diff > 0:
            result["options_trend"] = "😨 弱気（下落＆プット建玉増）"
        elif is_price_down and put_diff < 0:
            result["options_trend"] = "📈 下落は長続きせず反転上昇の可能性"
        else:
            result["options_trend"] = "❔ 判断困難"

    except Exception as e:
        result["options_trend"] = f"❌ エラー: {e}"

    return result